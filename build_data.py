"""Собрать auto_grp_data.js из файлов пластов. Запуск: python build_data.py"""
import os, json, csv, glob
import openpyxl

base = os.path.dirname(os.path.abspath(__file__))
NULL = 1e30


def discover_layers():
    """Находим папки пластов по trunk_info.csv (без хардкода кириллицы в путях)."""
    layers = {}
    for name in os.listdir(base):
        p = os.path.join(base, name)
        if not os.path.isdir(p):
            continue
        if not os.path.isfile(os.path.join(p, "trunk_info.csv")):
            continue
        nl = name.lower()
        if "2bs10" in nl or "2бс10" in nl:
            label = "2БС10"
        elif "bs9" in nl or "бс9" in nl:
            label = "БС9/1"
        else:
            label = name
        layers[label] = name
    return layers


def parse_grd(path, step=3):
    with open(path, encoding="utf-8", errors="ignore") as f:
        assert f.readline().strip() == "DSAA"
        nx, ny = map(int, f.readline().split())
        xmin, xmax = map(float, f.readline().split())
        ymin, ymax = map(float, f.readline().split())
        f.readline()
        vals = []
        for line in f:
            vals.extend(map(float, line.split()))
    dnx = (nx + step - 1) // step
    dny = (ny + step - 1) // step
    data = []
    for j in range(0, ny, step):
        for i in range(0, nx, step):
            z = vals[j * nx + i]
            data.append(round(z, 1) if z < NULL else None)
    return {
        "label": os.path.basename(path).replace(".grd", ""),
        "nx": dnx, "ny": dny,
        "xmin": xmin, "xmax": xmax, "ymin": ymin, "ymax": ymax,
        "step": step, "data": data,
    }


def parse_trunk(path):
    out = []
    with open(path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f, delimiter=";"):
            x = float(row["Координата Т1 Х, м"])
            y = float(row["Координата Т1 У, м"])
            x3s = (row.get("Координата Т3 Х, м") or "").strip()
            y3s = (row.get("Координата Т3 У, м") or "").strip()
            w = {"w": row["Скважина"].strip(), "x": x, "y": y}
            if x3s:
                w.update(x3=float(x3s), y3=float(y3s), hz=True)
            out.append(w)
    return out


def parse_blocks(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    blocks = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        bid = str(row[0]).strip()
        blocks.setdefault(bid, []).append([round(float(row[1]), 1), round(float(row[2]), 1)])
    wb.close()
    return blocks


def parse_mor(path):
    rows = []
    with open(path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f, delimiter=";"):
            d = row["Дата"]
            if not any(y in d for y in ("2024", "2025", "2026")):
                continue
            inj = float(str(row.get("Закачка за месяц, м3", 0)).replace(",", ".") or 0)
            liq = float(str(row.get("Добыча жидкости за месяц, т", 0)).replace(",", ".") or 0)
            oil = float(str(row.get("Добыча нефти за месяц, т", 0)).replace(",", ".") or 0)
            rows.append([row["Скважина"].strip(), d, round(inj, 1), round(liq, 1), round(oil, 1)])
    return rows


def norm_month_date(d):
    """01.MM.YYYY — как в МЭР и на картах."""
    parts = str(d).strip().split(".")
    if len(parts) == 3:
        return f"01.{parts[1]}.{parts[2]}"
    return d


def parse_tr(path):
    rows = []
    with open(path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f, delimiter=";"):
            d = row["Дата"]
            if not any(y in d for y in ("2024", "2025", "2026")):
                continue
            pzab_s = (row.get("Забойное давление (ТР), атм") or "").strip()
            if not pzab_s:
                continue
            pzab = float(pzab_s.replace(",", "."))
            rows.append([row["Скважина"].strip(), norm_month_date(d), round(pzab, 1)])
    return rows


def parse_history(path):
    from datetime import datetime, timedelta
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    data = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(data) < 3:
        return None
    hdr = data[0]
    wells = {}
    for c in range(1, len(hdr)):
        if not hdr[c]:
            continue
        wid = str(hdr[c]).strip().replace("Скв.", "").replace("скв.", "").strip()
        wells[wid] = c
    dates, series = [], {}
    for wid in wells:
        series[wid] = {"liq": [], "oil": []}
    for r in range(2, len(data)):
        row = data[r]
        d = row[0]
        if isinstance(d, (int, float)):
            d = datetime(1899, 12, 30) + timedelta(days=d)
        else:
            parts = str(d).split(".")
            if len(parts) != 3:
                continue
            d = datetime(int(parts[2]), int(parts[1]), int(parts[0]))
        if d.year > 2026 or (d.year == 2026 and d.month > 6):
            continue
        dates.append(f"{d.day:02d}.{d.month:02d}.{d.year}")
        for wid, col in wells.items():
            series[wid]["liq"].append(round(float(row[col] or 0), 1))
            series[wid]["oil"].append(round(float(row[col + 1] or 0), 1))
    return {"dates": dates, "wells": series}


def main():
    bundle = {"layers": {}}
    layers = discover_layers()
    if not layers:
        raise SystemExit("Не найдены папки пластов с trunk_info.csv")
    print("Пласты:", layers)
    for label, folder in layers.items():
        p = os.path.join(base, folder)
        maps = {}
        for fn in sorted(f for f in os.listdir(p) if f.lower().endswith(".grd")):
            full = os.path.join(p, fn)
            print(f"  GRD: {fn}")
            maps[fn] = parse_grd(full)
        tr_path = os.path.join(p, "tr_reports.csv")
        bundle["layers"][label] = {
            "maps": maps,
            "trunks": parse_trunk(os.path.join(p, "trunk_info.csv")),
            "blocks": parse_blocks(os.path.join(p, "block_coordinates.xlsx")),
            "mor": parse_mor(os.path.join(p, "mor_reports.csv")),
            "tr": parse_tr(tr_path) if os.path.isfile(tr_path) else [],
        }
        print(
            f"  {label}: {len(maps)} карт, {len(bundle['layers'][label]['trunks'])} стволов, "
            f"{len(bundle['layers'][label]['tr'])} записей ТР"
        )

    for fn in glob.glob(os.path.join(base, "*.xlsx")):
        if "(1)" in fn or os.path.getsize(fn) < 200000:
            continue
        print("history:", os.path.basename(fn))
        bundle["history"] = parse_history(fn)
        break

    out = os.path.join(base, "auto_grp_data.js")
    with open(out, "w", encoding="utf-8") as f:
        f.write("window.AUTO_GRP_DATA=")
        json.dump(bundle, f, ensure_ascii=False, separators=(",", ":"))
        f.write(";")

    print(f"OK: {out} ({os.path.getsize(out) // 1024} KB)")


if __name__ == "__main__":
    main()
