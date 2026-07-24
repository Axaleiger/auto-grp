#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Сборка JSON-пакета данных ЦДП для платформы.

Запуск в IDE / Jupyter:
  1. Скопируйте config.example.json → config.json
  2. Пропишите layer_id / field_id и период
  3. python export_data.py --config config.json
  4. Подключите JSON в каталоге платформы → «Данные»

Режимы:
  connection.mode = "local_csv"  — сбор из CSV/XLSX/GRD папок проекта (демо)
  connection.mode = "corp_api"   — корпоративный API (CorpApiClient)
"""
from __future__ import annotations

import argparse
import csv
import glob
import json
import os
import sys
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

try:
    import openpyxl
except ImportError:
    openpyxl = None

NULL = 1e30
ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
HERE = os.path.dirname(os.path.abspath(__file__))


# ─── утилиты парсинга (совместимы с build_data.py) ───────────────────────────

def parse_grd(path: str, step: int = 3) -> dict:
    with open(path, encoding="utf-8", errors="ignore") as f:
        assert f.readline().strip() == "DSAA"
        nx, ny = map(int, f.readline().split())
        xmin, xmax = map(float, f.readline().split())
        ymin, ymax = map(float, f.readline().split())
        f.readline()
        vals: List[float] = []
        for line in f:
            vals.extend(map(float, line.split()))
    dnx = (nx + step - 1) // step
    dny = (ny + step - 1) // step
    data = []
    for j in range(0, ny, step):
        for i in range(0, nx, step):
            z = vals[j * nx + i]
            data.append(round(z, 1) if z < NULL else None)
    return {
        "label": os.path.basename(path).replace(".grd", ""),
        "nx": dnx, "ny": dny,
        "xmin": xmin, "xmax": xmax, "ymin": ymin, "ymax": ymax,
        "step": step, "data": data,
    }


def parse_trunk(path: str) -> list:
    out = []
    with open(path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f, delimiter=";"):
            w = {
                "w": row["Скважина"].strip(),
                "x": float(row["Координата Т1 Х, м"]),
                "y": float(row["Координата Т1 У, м"]),
            }
            x3s = (row.get("Координата Т3 Х, м") or "").strip()
            y3s = (row.get("Координата Т3 У, м") or "").strip()
            if x3s:
                w.update(x3=float(x3s), y3=float(y3s), hz=True)
            out.append(w)
    return out


def parse_blocks(path: str) -> dict:
    if openpyxl is None:
        raise RuntimeError("Нужен openpyxl для block_coordinates.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    blocks: Dict[str, list] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        bid = str(row[0]).strip()
        blocks.setdefault(bid, []).append([round(float(row[1]), 1), round(float(row[2]), 1)])
    wb.close()
    return blocks


def parse_mor(path: str) -> list:
    rows = []
    with open(path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f, delimiter=";"):
            inj = float(str(row.get("Закачка за месяц, м3", 0)).replace(",", ".") or 0)
            liq = float(str(row.get("Добыча жидкости за месяц, т", 0)).replace(",", ".") or 0)
            oil = float(str(row.get("Добыча нефти за месяц, т", 0)).replace(",", ".") or 0)
            rows.append([row["Скважина"].strip(), row["Дата"], round(inj, 1), round(liq, 1), round(oil, 1)])
    return rows


def norm_month_date(d: str) -> str:
    parts = str(d).strip().split(".")
    if len(parts) == 3:
        return f"01.{parts[1]}.{parts[2]}"
    return d


def parse_tr(path: str) -> list:
    rows = []
    with open(path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f, delimiter=";"):
            pzab_s = (row.get("Забойное давление (ТР), атм") or "").strip()
            if not pzab_s:
                continue
            pzab = float(pzab_s.replace(",", "."))
            qliq_s = (row.get("Дебит жидкости, т/сут") or "").strip()
            qliq = float(qliq_s.replace(",", ".") or 0) if qliq_s else 0.0
            rows.append([
                row["Скважина"].strip(),
                norm_month_date(row["Дата"]),
                round(pzab, 1),
                round(qliq, 2),
            ])
    return rows


def parse_grp_ports_csv(path: str) -> dict:
    out = {}
    with open(path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f, delimiter=";"):
            w = (row.get("well") or row.get("Скважина") or "").strip()
            if not w:
                continue
            out[w] = {
                "w": w,
                "ports": int(float(row.get("ports") or row.get("Порты") or 0)),
                "az": float(str(row.get("azimuth") or row.get("Азимут") or 0).replace(",", ".")),
                "L": float(str(row.get("gs_length") or row.get("L") or 0).replace(",", ".")),
            }
    return out


def parse_history(path: str) -> Optional[dict]:
    if openpyxl is None:
        return None
    from datetime import datetime, timedelta
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    data = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(data) < 3:
        return None
    hdr = data[0]
    wells = {}
    for c in range(1, len(hdr)):
        if not hdr[c]:
            continue
        wid = str(hdr[c]).strip().replace("Скв.", "").replace("скв.", "").strip()
        wells[wid] = c
    dates, series = [], {}
    for wid in wells:
        series[wid] = {"liq": [], "oil": []}
    for r in range(2, len(data)):
        row = data[r]
        d = row[0]
        if isinstance(d, (int, float)):
            d = datetime(1899, 12, 30) + timedelta(days=d)
        else:
            parts = str(d).split(".")
            if len(parts) != 3:
                continue
            d = datetime(int(parts[2]), int(parts[1]), int(parts[0]))
        if d.year > 2026 or (d.year == 2026 and d.month > 6):
            continue
        dates.append(f"{d.day:02d}.{d.month:02d}.{d.year}")
        for wid, col in wells.items():
            series[wid]["liq"].append(round(float(row[col] or 0), 1))
            series[wid]["oil"].append(round(float(row[col + 1] or 0), 1))
    return {"dates": dates, "wells": series}


# ─── источники данных ────────────────────────────────────────────────────────────

class DataSourceClient:
    """Интерфейс источника данных. Реализуйте методы под корпоративный API/SQL."""

    def fetch_trunks(self, layer_id: str) -> list:
        raise NotImplementedError

    def fetch_blocks(self, layer_id: str) -> dict:
        raise NotImplementedError

    def fetch_mor(self, layer_id: str, date_from: str, date_to: str) -> list:
        raise NotImplementedError

    def fetch_tr(self, layer_id: str, date_from: str, date_to: str) -> list:
        raise NotImplementedError

    def fetch_maps(self, layer_id: str, date_from: str, date_to: str) -> dict:
        raise NotImplementedError

    def fetch_grp_ports(self, layer_id: str) -> dict:
        raise NotImplementedError

    def fetch_history(self, field_id: str, date_from: str, date_to: str) -> Optional[dict]:
        raise NotImplementedError

    def fetch_startup(self, layer_id: str) -> dict:
        raise NotImplementedError


class LocalCsvClient(DataSourceClient):
    """Демо-адаптер: читает локальные папки пластов (как build_data.py)."""

    def __init__(self, root: str, layer_folders: Dict[str, str]):
        self.root = root
        self.layer_folders = layer_folders  # label -> folder name

    def _folder(self, label: str) -> str:
        folder = self.layer_folders.get(label)
        if not folder:
            raise FileNotFoundError(f"Не задана local_folder для пласта {label}")
        path = os.path.join(self.root, folder)
        if not os.path.isdir(path):
            raise FileNotFoundError(path)
        return path

    def fetch_trunks(self, layer_id: str) -> list:
        # layer_id здесь = label для local режима
        return parse_trunk(os.path.join(self._folder(layer_id), "trunk_info.csv"))

    def fetch_blocks(self, layer_id: str) -> dict:
        p = os.path.join(self._folder(layer_id), "block_coordinates.xlsx")
        return parse_blocks(p) if os.path.isfile(p) else {}

    def fetch_mor(self, layer_id: str, date_from: str, date_to: str) -> list:
        return parse_mor(os.path.join(self._folder(layer_id), "mor_reports.csv"))

    def fetch_tr(self, layer_id: str, date_from: str, date_to: str) -> list:
        p = os.path.join(self._folder(layer_id), "tr_reports.csv")
        return parse_tr(p) if os.path.isfile(p) else []

    def fetch_maps(self, layer_id: str, date_from: str, date_to: str) -> dict:
        folder = self._folder(layer_id)
        maps = {}
        for fn in sorted(f for f in os.listdir(folder) if f.lower().endswith(".grd")):
            maps[fn] = parse_grd(os.path.join(folder, fn))
        return maps

    def fetch_grp_ports(self, layer_id: str) -> dict:
        for candidate in (
            os.path.join(self.root, "roma_grp_ports.csv"),
            os.path.join(self._folder(layer_id), "roma_grp_ports.csv"),
        ):
            if os.path.isfile(candidate):
                return parse_grp_ports_csv(candidate)
        return {}

    def fetch_history(self, field_id: str, date_from: str, date_to: str) -> Optional[dict]:
        for fn in glob.glob(os.path.join(self.root, "*.xlsx")):
            if "(1)" in fn or os.path.getsize(fn) < 200000:
                continue
            return parse_history(fn)
        return None

    def fetch_startup(self, layer_id: str) -> dict:
        # Дефолты; подставьте из корпоративного хранилища PVT/ГФ/глубины по layer_id
        return {
            "kh": 10, "h": 8, "phi": 0.2, "ik2": 0.1,
            "mu_o": 1.5, "mu_w": 0.5, "Bo": 1.2, "Bw": 1,
            "ro_o": 0.85, "Pb": 80, "L": 400, "re": 250, "rw": 0.1,
            "Pi": 180, "Pwf": 60, "fw": 40, "Gf": 80,
            "Hform": 2500, "Hpump": 2200, "FracCount": 0, "xf": 50,
        }


class CorpApiClient(DataSourceClient):
    """
    Каркас корпоративного API. Замените URL/параметры на реальные эндпоинты.

    Ожидаемая модель:
      GET /layers/{layer_id}/wells
      GET /layers/{layer_id}/blocks
      GET /layers/{layer_id}/mer?from=&to=
      GET /layers/{layer_id}/tr?from=&to=
      GET /layers/{layer_id}/pressure-maps?from=&to=
      GET /layers/{layer_id}/frac-ports
      GET /fields/{field_id}/production-history?from=&to=
      GET /layers/{layer_id}/startup-params
    """

    def __init__(self, base_url: str, token: str, timeout: int = 120):
        self.base_url = base_url.rstrip("/")
        self.token = token
        self.timeout = timeout
        try:
            import urllib.request
            self._urllib = urllib.request
        except ImportError as e:
            raise RuntimeError("urllib недоступен") from e

    def _get(self, path: str, params: Optional[dict] = None) -> Any:
        from urllib.parse import urlencode
        url = self.base_url + path
        if params:
            url += "?" + urlencode({k: v for k, v in params.items() if v is not None})
        req = self._urllib.Request(url, headers={
            "Authorization": f"Bearer {self.token}",
            "Accept": "application/json",
        })
        with self._urllib.urlopen(req, timeout=self.timeout) as resp:
            return json.loads(resp.read().decode("utf-8"))

    def fetch_trunks(self, layer_id: str) -> list:
        # TODO: привести ответ API к [{w,x,y,x3?,y3?,hz?}]
        data = self._get(f"/layers/{layer_id}/wells")
        return data if isinstance(data, list) else data.get("items", [])

    def fetch_blocks(self, layer_id: str) -> dict:
        data = self._get(f"/layers/{layer_id}/blocks")
        return data if isinstance(data, dict) else data.get("blocks", {})

    def fetch_mor(self, layer_id: str, date_from: str, date_to: str) -> list:
        data = self._get(f"/layers/{layer_id}/mer", {"from": date_from, "to": date_to})
        return data if isinstance(data, list) else data.get("rows", [])

    def fetch_tr(self, layer_id: str, date_from: str, date_to: str) -> list:
        data = self._get(f"/layers/{layer_id}/tr", {"from": date_from, "to": date_to})
        return data if isinstance(data, list) else data.get("rows", [])

    def fetch_maps(self, layer_id: str, date_from: str, date_to: str) -> dict:
        data = self._get(f"/layers/{layer_id}/pressure-maps", {"from": date_from, "to": date_to})
        return data if isinstance(data, dict) else data.get("maps", {})

    def fetch_grp_ports(self, layer_id: str) -> dict:
        data = self._get(f"/layers/{layer_id}/frac-ports")
        return data if isinstance(data, dict) else data.get("ports", {})

    def fetch_history(self, field_id: str, date_from: str, date_to: str) -> Optional[dict]:
        data = self._get(f"/fields/{field_id}/production-history", {"from": date_from, "to": date_to})
        return data

    def fetch_startup(self, layer_id: str) -> dict:
        data = self._get(f"/layers/{layer_id}/startup-params")
        return data if isinstance(data, dict) else {}


# ─── сборка экспорта ─────────────────────────────────────────────────────────

def build_export(cfg: dict, client: DataSourceClient) -> dict:
    datasets = set(cfg.get("datasets") or [])
    period = cfg.get("period") or {}
    date_from = period.get("date_from")
    date_to = period.get("date_to")
    meta = dict(cfg.get("meta") or {})

    export: Dict[str, Any] = {
        "schema": "cdp-platform-export",
        "version": "1.0",
        "exported_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source": {
            "system": "corp_api" if cfg.get("connection", {}).get("mode") == "corp_api" else "local_csv",
            "config_meta": {k: meta.get(k) for k in ("do", "field", "field_id")},
        },
        "meta": meta,
        "request": {
            "period": period,
            "datasets": sorted(datasets),
            "layers": [
                {"label": L.get("label"), "layer_id": L.get("layer_id")}
                for L in cfg.get("layers") or []
            ],
        },
        "layers": {},
    }

    for L in cfg.get("layers") or []:
        label = L["label"]
        # В local_csv режиме ключ запросов = label; в corp_api = layer_id
        key = label if isinstance(client, LocalCsvClient) else L["layer_id"]
        layer_obj: Dict[str, Any] = {
            "layer_id": L.get("layer_id"),
            "label": label,
            "field_id": meta.get("field_id"),
            "field": meta.get("field"),
        }
        print(f"• {label} ({key})")
        if "trunks" in datasets:
            layer_obj["trunks"] = client.fetch_trunks(key)
            print(f"  trunks: {len(layer_obj['trunks'])}")
        if "blocks" in datasets:
            layer_obj["blocks"] = client.fetch_blocks(key)
            print(f"  blocks: {len(layer_obj['blocks'])}")
        if "mor" in datasets:
            layer_obj["mor"] = client.fetch_mor(key, date_from, date_to)
            print(f"  mor: {len(layer_obj['mor'])}")
        if "tr" in datasets:
            layer_obj["tr"] = client.fetch_tr(key, date_from, date_to)
            print(f"  tr: {len(layer_obj['tr'])}")
        if "maps" in datasets:
            layer_obj["maps"] = client.fetch_maps(key, date_from, date_to)
            print(f"  maps: {len(layer_obj['maps'])}")
        if "grp_ports" in datasets:
            layer_obj["grp_ports"] = client.fetch_grp_ports(key)
            print(f"  grp_ports: {len(layer_obj['grp_ports'])}")
        if "startup" in datasets:
            layer_obj["startup"] = client.fetch_startup(key)
        export["layers"][label] = layer_obj

    if "history" in datasets:
        field_id = meta.get("field_id") or meta.get("field") or ""
        hist = client.fetch_history(field_id, date_from, date_to)
        if hist:
            export["history"] = hist
            print(f"• history: {len(hist.get('dates', []))} дат, {len(hist.get('wells', {}))} скв.")
        else:
            print("• history: нет данных")

    return export


def to_auto_grp_bundle(export: dict) -> dict:
    """Преобразование экспорта → формат AUTO_GRP_DATA (+ grp_ports на верхнем уровне)."""
    layers = {}
    grp_all = {}
    for label, L in (export.get("layers") or {}).items():
        layers[label] = {
            "maps": L.get("maps") or {},
            "trunks": L.get("trunks") or [],
            "blocks": L.get("blocks") or {},
            "mor": L.get("mor") or [],
            "tr": L.get("tr") or [],
        }
        for w, port in (L.get("grp_ports") or {}).items():
            grp_all[w] = port
    bundle: Dict[str, Any] = {"layers": layers}
    if export.get("history"):
        bundle["history"] = export["history"]
    if grp_all:
        bundle["grp_ports"] = grp_all
    # startup defaults per layer
    startup = {label: L["startup"] for label, L in (export.get("layers") or {}).items() if L.get("startup")}
    if startup:
        bundle["startup"] = startup
    return bundle


def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser(description="Выгрузка данных ЦДП → JSON")
    ap.add_argument("--config", default=os.path.join(HERE, "config.json"),
                    help="Путь к config.json (по умолчанию data_export/config.json)")
    args = ap.parse_args(argv)

    cfg_path = args.config
    if not os.path.isfile(cfg_path):
        example = os.path.join(HERE, "config.example.json")
        print(f"Нет {cfg_path}. Скопируйте {example} → config.json и заполните layer_id.")
        return 1

    with open(cfg_path, encoding="utf-8") as f:
        cfg = json.load(f)

    mode = (cfg.get("connection") or {}).get("mode", "local_csv")
    if mode == "local_csv":
        folders = {L["label"]: L.get("local_folder") for L in cfg.get("layers") or []}
        client: DataSourceClient = LocalCsvClient(ROOT, folders)
    elif mode == "corp_api":
        api_cfg = (cfg.get("connection") or {}).get("api") or {}
        token = os.environ.get(api_cfg.get("token_env") or "CDP_DATA_TOKEN", "")
        if not token:
            print("Задайте токен в переменной окружения", api_cfg.get("token_env") or "CDP_DATA_TOKEN")
            return 1
        client = CorpApiClient(api_cfg["base_url"], token, int(api_cfg.get("timeout_sec") or 120))
    else:
        print("Неизвестный connection.mode:", mode)
        return 1

    export = build_export(cfg, client)
    out_cfg = cfg.get("output") or {}
    json_path = out_cfg.get("json_path") or os.path.join(HERE, "out", "cdp_export.json")
    if not os.path.isabs(json_path):
        json_path = os.path.join(ROOT, json_path)
    os.makedirs(os.path.dirname(json_path), exist_ok=True)

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(export, f, ensure_ascii=False, separators=(",", ":"))
    print(f"OK JSON: {json_path} ({os.path.getsize(json_path) // 1024} KB)")

    if out_cfg.get("also_write_js"):
        js_path = out_cfg.get("js_path") or "auto_grp_data.js"
        if not os.path.isabs(js_path):
            js_path = os.path.join(ROOT, js_path)
        bundle = to_auto_grp_bundle(export)
        with open(js_path, "w", encoding="utf-8") as f:
            f.write("window.AUTO_GRP_DATA=")
            json.dump(bundle, f, ensure_ascii=False, separators=(",", ":"))
            f.write(";")
        print(f"OK JS: {js_path}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
